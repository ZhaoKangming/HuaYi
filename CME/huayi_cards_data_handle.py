# -*- coding:utf-8 -*-
'''
Function：处理华医学术卡的数据表，生成数据周报
Author: ZhaoKangming
E-mail: zhaokm0@gmail.com
Version: V0.1
'''
import urllib.request
import io
import requests
import shutil
import os
from sys import intern
import win32com.client as win32
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, NamedStyle
import sys
import time
import re


''' 
------------------- 提前的准备工作 -----------------------
1. 在《企业投放统计》表格中检查卡类型是否完全，购卡数量是否更新


------------------- 之后的收尾工作 -----------------------
1.因用openpyxl会使得表格的样式变化，请将表格单独复制，并用Chart_Data进行数据替换
2.更新一页图表中的周报完成日期与时间的更新
3.使用格式刷来将新插入的行的样式使之一致
4.箱图和地图复制粘贴为图片，以防低版本office或者wps不兼容，显示失败
'''

# 全局变量的定义及赋值
workspace_path: str = os.path.dirname(os.path.realpath(__file__))
today_date: str = str(datetime.date.today()).replace("-", "").replace('2019','19')
step_numb: int = 0


def download_data_xls() -> list:
    '''
    【功能】从数据统计后台中下载相应的华医学术卡数据原始记录表
    '''
    card_data_dict: dict = {'kind': 'card_yaoqi'}

    #-------------------------- 模拟登陆 --------------------------
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf8')  # 改变标准输出的默认编码
    data = {'login_name': 'yaoqi', 'login_pwd': 1}  # 登录时需要POST的数据
    login_url: str = 'http://192.168.1.240:8122/Login/Login'  # 登录时表单提交到的地址
    session = requests.Session()
    resp = session.post(login_url, data)

    #-------------------------- 请求数据 --------------------------
    headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
                'Accept-Encoding': 'gzip, deflate',
                'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
                'Cache-Control': 'max-age=0',
                'Connection': 'keep-alive',
                'Content-Type': 'application/x-www-form-urlencoded',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36'}
    download_url: str = 'http://192.168.1.240:8122/Common/ToExcel'  # 下载POST地址

    download_state_list: list = []
    card_download_resp = session.post(download_url, card_data_dict, headers = headers)
    card_file_path: str = os.path.join(workspace_path, 'data', f'华医学术卡原始数据-{today_date}.xls')
    with open(card_file_path, "wb") as card_downloaded_file:
        card_downloaded_file.write(card_download_resp.content)
    download_state_list.append(card_download_resp.status_code)
    download_state_list.append(card_file_path)

    return download_state_list


def xls_to_xlsx(xls_path: str) -> str:
    '''
    【功能】将 xls 文件转化为 xlsx 文件
    :param xls_path: xls文件的路径
    ''' 
    # 文件格式转化
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(xls_path)
    xlsx_path: str = xls_path + 'x'
    wb.SaveAs(xlsx_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    # 源文件处理与新文件输出
    os.remove(xls_path)
    return xlsx_path


def only_chinese(content: str) -> str:
    '''
    【功能】将传入的文本仅保留中文
    '''
    # 处理前进行相关的处理，包括转换成Unicode等
    pattern = re.compile('[^\u4e00-\u9fa50-9]')  # 中文的编码范围是：\u4e00到\u9fa5
    zh_str: str = "".join(pattern.split(content))
    return zh_str



def statistic_data():
    '''
    【功能】进行华医学术卡数据统计分析
    '''
    global step_numb
    # ------------------- 获取模板与数据文件并备份 -------------------
    # 载入表格
    data_xlsx_path: str = os.path.join(workspace_path, 'data', f'华医学术卡原始数据-{today_date}.xlsx')
    data_wb = load_workbook(data_xlsx_path)

    template_xlsx_path: str = os.path.join(workspace_path, '【数据模板】华医网学术卡数据周报.xlsx')
    template_wb = load_workbook(template_xlsx_path)
    chart_data_sht = template_wb['Chart_Data']
    prov_sht = template_wb['省份分布']
    card_sht = template_wb['卡类状况']
    cpy_sht = template_wb['企业投放统计']


    # 备份并重命名《学习记录》原始数据表
    data_sht = data_wb.copy_worksheet(data_wb['sheet0'])
    data_wb['sheet0'].title = 'backup'
    data_sht.title = 'data'
    data_last_row: int = data_sht.max_row
    sum_card_numb_now: int = data_last_row - 1       # 当前全国总绑卡数

    step_numb += 1
    print(f'【STEP-{step_numb}】表格载入与数据备份\n\t\t[OK] --> 已经成功备份数据表格!\n')

    # ------------------- 数据清洗 -------------------
    # 删除多余的列
    spare_cols_list: list = [15, 12, 10, 9, 6, 4, 3, 1]
    for spare_col_numb in spare_cols_list:
        data_sht.delete_cols(spare_col_numb)

    # 插入新列并获取绑卡的月份，小时
    data_sht.insert_cols(3, 2)
    for i in range(2, data_last_row + 1):
        data_sht.cell(i, 3).value = str(data_sht.cell(i, 2).value)[5:7]  # 月份
        data_sht.cell(i, 4).value = str(data_sht.cell(i, 2).value)[11:13] # 小时

    # 设置列名
    col_name_list: list = ['卡类型','绑定时间','月份','小时','省份','城市','单位级别','职称','专业']
    for i in range(len(col_name_list)):
        data_sht.cell(1, i+1).value = col_name_list[i]

    # 数据清洗
    content_col_dict: dict = {'省份':'5' , '城市':'6' , '单位级别':'7' , '职称':'8' , '专业':'9'}
    outlier_list: list = ['', '其它', '-请选择-', 'NULL', 'null','请选择']
    city_outlier_list: list = ['省直属单位', '省直辖县级行政单位', '省直辖县级行政区划']
    #TODO:检查dict是否有重复的列值，有问题则报错
    for i in range(2, data_last_row + 1):
        # 统一遍历清洗
        for k,v in content_col_dict.items():
            if not v == '':
                if not data_sht.cell(i, int(v)).value and data_sht.cell(i, int(v)).value != 0:
                    data_sht.cell(i, int(v)).value = '其他'
                else:
                    data_sht.cell(i, int(v)).value = only_chinese(data_sht.cell(i, int(v)).value) # 仅保留中文文本
                    if data_sht.cell(i, int(v)).value in outlier_list:
                        data_sht.cell(i, int(v)).value = '其他'
                    if k == '城市' and data_sht.cell(i, int(v)).value in city_outlier_list:
                        data_sht.cell(i, int(v)).value = data_sht.cell(i, int(content_col_dict['省份'])).value + data_sht.cell(i, int(v)).value
                    if k == '省份' and data_sht.cell(i, int(v)).value == '黑龙江森林工业总局卫生局':
                        data_sht.cell(i, int(v)).value = '黑龙江森工'

    step_numb += 1
    print(f'【STEP-{step_numb}】数据清洗\n\t\t[OK] --> 已经完成数据清洗!\n')

    # ------------------- 数据统计 -------------------
    prov_card_dict: dict = {}
    prov_card: str = ''
    orig_prov_dict: dict = {}   # 对省份不进行合并
    prov_dict: dict = {"广东省" : 0, "四川省" : 0, "海南省" : 0, "山西省" : 0, "江苏省" : 0, "山东省" : 0, "北京市" : 0, "江西省" : 0, "重庆市" : 0, "湖南省" : 0, "河南省" : 0, "福建省" : 0, "贵州省" : 0, "黑龙江省" : 0, "安徽省" : 0, "新疆维吾尔自治区" : 0, "吉林省" : 0, "河北省" : 0, "广西壮族自治区" : 0, "辽宁省" : 0, "浙江省" : 0, "陕西省" : 0, "宁夏回族自治区" : 0, "湖北省" : 0, "甘肃省" : 0, "青海省" : 0, "西藏自治区" : 0, "天津市" : 0, "上海市" : 0, "云南省" : 0, "内蒙古自治区" : 0, "香港特别行政区" : 0, "澳门特别行政区" : 0, "台湾省" : 0, "其他" : 0} 
    pro: str = ''
    card_dict: dict = {}
    city_dict: dict = {}
    hour_dict: dict = {"00": 0, "01": 0, "02": 0, "03": 0, "04": 0, "05": 0, "06": 0, "07": 0, "08": 0, "09": 0, "10": 0, "11": 0, "12": 0, "13": 0, "14": 0, "15": 0, "16": 0, "17": 0, "18": 0, "19": 0, "20": 0, "21": 0, "22": 0, "23": 0}
    month_dict: dict = {"02" : 0, "03" : 0, "04" : 0, "05" : 0, "06" : 0, "07" : 0, "08" : 0, "09" : 0, "10" : 0, "11" : 0, "12" : 0}
    hosp_dict: dict = {"三甲" : 0, "三乙" : 0, "二甲" : 0, "二乙" : 0, "一甲" : 0, "一乙" : 0, "其他" : 0}


    for i in range(2, data_last_row + 1):
        prov_card = data_sht.cell(i, 5).value + '#' + data_sht.cell(i, 1).value
        # 省份-卡类型字典统计
        prov_card_dict.setdefault(prov_card, 0)
        prov_card_dict[prov_card] += 1

        orig_prov_dict.setdefault(data_sht.cell(i, 5).value, 0)
        orig_prov_dict[data_sht.cell(i, 5).value] += 1

        # 省份统计
        if '新疆' in data_sht.cell(i, 5).value:
            pro = '新疆维吾尔自治区'           # 发卡区域合并：新疆生产建设兵团
        elif '黑龙江' in data_sht.cell(i, 5).value:
            pro = '黑龙江省'                    # 发卡区域合并：黑龙江森林工业总局卫生局、黑龙江农垦、黑龙江森工
        else:
            pro = data_sht.cell(i, 5).value
        prov_dict[pro] += 1

        # 城市统计
        city_dict.setdefault(data_sht.cell(i, 6).value, 0)
        city_dict[data_sht.cell(i, 6).value] += 1

        # 卡类型统计
        card_dict.setdefault(data_sht.cell(i, 1).value, 0)
        card_dict[data_sht.cell(i, 1).value] += 1

        hour_dict[data_sht.cell(i, 4).value] += 1      # 小时统计        
        month_dict[data_sht.cell(i, 3).value] += 1      # 月份统计
        hosp_dict[data_sht.cell(i, 7).value] += 1      # 医院级别统计

    step_numb += 1
    print(f'【STEP-{step_numb}】数据统计\n\t\t[OK] --> 已经完成数据统计!\n')

    # ------------------- 其他信息的读取 -------------------
    #.......... 省份限制类信息的读取 ............
    prov_limit_dict: dict = {}      # 省份的限制信息字典
    for i in range(2, template_wb['Prov_Limit'].max_row + 1):
        prov_limit_dict[template_wb['Prov_Limit'].cell(i,1).value] = [template_wb['Prov_Limit'].cell(i,2).value, template_wb['Prov_Limit'].cell(i,3).value]
    
    #.......... 企业投放类信息的读取 ............
    cpy_last_row: int = cpy_sht.max_row  # 企业投放统计表的最后一行行号
    sold_card_numb: int = 0             # 已经售出的卡数量总和
    card_info_dict: dict = {}           # 记录卡类型的所属企业和购卡数量信息                   
    for i in range(2, cpy_last_row):
        sold_card_numb += cpy_sht.cell(i,7).value
        card_info_dict[cpy_sht.cell(i, 5).value] = [cpy_sht.cell(i, 1).value, cpy_sht.cell(i,7).value]

    #.......... 其他类信息的读取 ............
    delta = datetime.datetime.now() - datetime.datetime.strptime('2019-02-01', '%Y-%m-%d')  # 今天与190201相距多少天
    avg_week_card: int = int(sum_card_numb_now/int(delta.days))*7  # 平均周绑卡数量

    step_numb += 1
    print(f'【STEP-{step_numb}】其他数据读取\n\t\t[OK] --> 已经完成相关辅助数据的读取!\n')


    # ------------------- 数据写入 -------------------
    # ............《省份分布表》............
    prov_sht.insert_cols(9)
    prov_sht['I1'].value = today_date
    prov_last_col: int = prov_sht.max_column
    # 插入新的省份与卡类型
    for k,v in prov_card_dict.items():
        prov_last_row: int = prov_sht.max_row
        is_new_item: bool = True
        is_new_prov: bool = True
        # 处理已经存在的省份-卡类型
        for i in range(2, prov_last_row):
            current_prov_card: str = prov_sht.cell(i,1).value + '#' + prov_sht.cell(i,7).value
            if current_prov_card == k:
                prov_sht.cell(i,9).value = v
                is_new_item = False
                is_new_prov = False
                break
        # 处理新增的省份-卡类型
        if is_new_item == True:
            for i in range(2, prov_last_row):
                if prov_sht.cell(i, 1).value == k.split('#')[0]:
                    is_new_prov = False
                    # 处理当前投放机构首次出现绑卡的状况
                    if  prov_sht.cell(i, 7).value == '—':
                        prov_sht.cell(i, 7).value = k.split('#')[1]
                        prov_sht.cell(i, 9).value = v
                    # 处理当前投放机构之前有绑卡，本次新增卡类型的状况
                    else:
                        prov_sht.insert_rows(i)
                        prov_sht.cell(i, 1).value = k.split('#')[0]
                        prov_sht.cell(i, 7).value = k.split('#')[1]
                        prov_sht.cell(i, 9).value = v
                    for j in range(10, prov_last_col + 1):
                        prov_sht.cell(i, j).value = 0
                    break
        # 处理新增的投放机构
        if is_new_prov == True:
            temp_prov = k.split('#')[0]
            print(f'>>>>>>>> 出现了新的发卡机构【{temp_prov}】，请注意 <<<<<<<<\n')
            prov_sht.insert_rows(prov_last_row-1)
            prov_sht.cell(prov_last_row-1, 1).value = k.split('#')[0]
            prov_sht.cell(prov_last_row-1, 7).value = k.split('#')[1]
            prov_sht.cell(prov_last_row-1, 9).value = v
            for j in range(10, prov_last_col + 1):
                prov_sht.cell(prov_last_row-1, j).value = 0

    prov_last_row = prov_sht.max_row

    for i in range(2, prov_last_row):
        # 将之前有过绑卡记录但是现在为0的以及至今从未发出卡的投放机构用 0来补全
        if not prov_sht.cell(i, 1).value + '#' + prov_sht.cell(i, 7).value in prov_card_dict.keys():
            prov_sht.cell(i, 9).value = 0

        if prov_sht.cell(i, 1).value in prov_limit_dict.keys():
            prov_sht.cell(i, 2).value = prov_limit_dict[prov_sht.cell(i, 1).value][0]         # 写入省份的投放限制备注
            prov_sht.cell(i, 3).value = prov_limit_dict[prov_sht.cell(i, 1).value][1]         # 写入省份的限制数
        
        if prov_sht.cell(i, 1).value in orig_prov_dict.keys():                                # 计算省份已发卡数
            prov_sht.cell(i, 4).value = orig_prov_dict[prov_sht.cell(i, 1).value]
        else:
            prov_sht.cell(i, 4).value = 0
        if prov_sht.cell(i, 3).value == '—':
            prov_sht.cell(i, 5).value = 0                                                     # 计算卡的剩余量
            prov_sht.cell(i, 6).value = '超出！'
        else:
            prov_sht.cell(i, 5).value = prov_sht.cell(i, 3).value - prov_sht.cell(i, 4).value     # 计算卡的剩余量
            prov_sht.cell(i, 6).value = prov_sht.cell(i, 4).value / prov_sht.cell(i, 3).value # 计算投放进度
        if prov_sht.cell(i, 6).value != '超出！' and prov_sht.cell(i, 6).value > 100:
            prov_sht.cell(i, 6).value = '超出！'
        prov_sht.cell(i, 8).value = prov_sht.cell(i, 9).value - prov_sht.cell(i, 10).value    # 计算本周增加数

    # 计算最后一行的总结列
    prov_sht.cell(prov_last_row, 3).value  = prov_limit_dict['总计'][1]
    prov_sht.cell(prov_last_row, 4).value = sum_card_numb_now
    prov_sht.cell(prov_last_row, 5).value = prov_limit_dict['总计'][1] - sum_card_numb_now
    prov_sht.cell(prov_last_row, 6).value = prov_sht.cell(prov_last_row, 4).value / prov_sht.cell(prov_last_row, 3).value
    prov_sht.cell(prov_last_row, 8).value = sum_card_numb_now - prov_sht.cell(prov_last_row, 10).value
    prov_sht.cell(prov_last_row, 9).value = sum_card_numb_now



    # ............《卡类状况表》............
    card_sht.insert_cols(9)
    card_sht['I1'].value = today_date
    card_last_col: int = card_sht.max_column
    card_last_row: int = card_sht.max_row

    for k, v in prov_card_dict.items():
        card_last_row: int = card_sht.max_row
        is_new_item: bool = True
        is_new_cardtype: bool = True
        # 处理已经存在的 卡类型-省份
        for i in range(2, card_last_row):
            current_prov_card: str = card_sht.cell(i, 7).value + '#' + card_sht.cell(i, 1).value
            if current_prov_card == k:
                card_sht.cell(i, 9).value = v
                is_new_item = False
                is_new_cardtype = False
                break
        # 处理新增的 卡类型-省份
        if is_new_item == True:
            for i in range(2, card_last_row):
                if card_sht.cell(i, 1).value == k.split('#')[1]:
                    is_new_cardtype = False
                    # 处理当前卡类型首次出现投放区域的状况
                    if card_sht.cell(i, 7).value == '—':
                        card_sht.cell(i, 7).value = k.split('#')[0]
                        card_sht.cell(i, 9).value = v
                    # 处理当前卡类型之前有绑卡，本次新增投放机构的状况
                    else:
                        card_sht.insert_rows(i)
                        card_sht.cell(i, 1).value = k.split('#')[1]
                        card_sht.cell(i, 7).value = k.split('#')[0]
                        card_sht.cell(i, 9).value = v
                    for j in range(10, card_last_col + 1):
                        card_sht.cell(i, j).value = 0
                    break
        # 处理新增的卡类型
        if is_new_cardtype == True:
            temp_cardtype = k.split('#')[1]
            print(f'>>>>>>>> 出现了新的卡类型【{temp_cardtype}】，请注意 <<<<<<<<\n')
            card_sht.insert_rows(card_last_row-1)
            card_sht.cell(card_last_row-1, 1).value = k.split('#')[1]
            card_sht.cell(card_last_row-1, 7).value = k.split('#')[0]
            card_sht.cell(card_last_row-1, 9).value = v
            for j in range(10, card_last_col + 1):
                card_sht.cell(card_last_row-1, j).value = 0

        card_last_row = card_sht.max_row
    
    card_last_row = card_sht.max_row

    for i in range(2, card_last_row):
        # 将之前有过绑卡记录但是现在为0的以及至今从未发出卡的卡类型用 0来补全
        if not card_sht.cell(i, 7).value + '#' + card_sht.cell(i, 1).value in prov_card_dict.keys():
            card_sht.cell(i, 9).value = 0

        # if card_sht.cell(i, 1).value in card_info_dict.keys():
        
        card_sht.cell(i, 2).value = card_info_dict[card_sht.cell(i, 1).value][0]         # 写入卡类型所属的企业
        card_sht.cell(i, 3).value = card_info_dict[card_sht.cell(i, 1).value][1]         # 写入卡类型的购卡数量

        # 计算卡类型总已发卡数
        if card_sht.cell(i, 1).value in card_dict.keys():
            card_sht.cell(i, 5).value = card_dict[card_sht.cell(i, 1).value]
        else:
            card_sht.cell(i, 5).value = 0
        
        card_sht.cell(i, 4).value = card_sht.cell(i, 5).value / card_sht.cell(i, 3).value       # 本卡投放进度
        card_sht.cell(i, 6).value = card_sht.cell(i, 5).value / sum_card_numb_now               # 占所有卡的投放比例
        card_sht.cell(i, 8).value = card_sht.cell(i, 9).value - card_sht.cell(i, 10).value      # 本周增长数

    # 计算最后一行的总结列
    card_sht.cell(card_last_row, 3).value = sold_card_numb                          # 华医售卡总数量
    card_sht.cell(card_last_row, 4).value = sum_card_numb_now / sold_card_numb      # 所有卡的投放总进度
    card_sht.cell(card_last_row, 5).value = sum_card_numb_now                       # 所有卡总绑卡数
    card_sht.cell(card_last_row, 6).value = 1                                       #占所有已投放卡的比例
    card_sht.cell(card_last_row, 8).value = sum_card_numb_now - card_sht.cell(card_last_row, 10).value
    card_sht.cell(card_last_row, 9).value = sum_card_numb_now

    #............《企业投放统计表》............
    cpy_sht.insert_cols(10)
    cpy_sht['J1'].value = today_date
    cpy_dict: dict = {}
    
    for k,v in card_info_dict.items():
        cpy_dict.setdefault(v[0],[0,0])             # 如果key第一次出现，设置该key的值为列表 [企业总购卡数为0，企业总绑卡数为0]
        cpy_dict[v[0]][0] += v[1]                   # 企业累积购卡数
        if k in card_dict.keys():
            cpy_dict[v[0]][1] += card_dict[k]       # 企业累积绑卡数 


    for i in range(2, cpy_last_row):
        # 计算企业的相关数据
        cpy_sht.cell(i,2).value = cpy_dict[cpy_sht.cell(i,1).value][0]                # 企业总购卡数
        cpy_sht.cell(i,3).value = cpy_dict[cpy_sht.cell(i,1).value][1]                # 企业总绑卡数
        cpy_sht.cell(i,4).value = cpy_sht.cell(i,3).value / cpy_sht.cell(i,2).value   # 企业投放进度

        # 计算卡类的相关数据
        if cpy_sht.cell(i,5).value in card_dict.keys():
            cpy_sht.cell(i,10).value = card_dict[cpy_sht.cell(i,5).value]
        else:
            cpy_sht.cell(i, 10).value = 0
        cpy_sht.cell(i, 8).value = cpy_sht.cell(i, 10).value / cpy_sht.cell(i,7).value          # 计算卡的投放进度
        cpy_sht.cell(i, 9).value = cpy_sht.cell(i, 10).value - cpy_sht.cell(i, 11).value        # 计算周增长数

    # 计算最后一列的相关汇总性数据
    cpy_sht.cell(cpy_last_row, 2).value = sold_card_numb                                               # 所有企业的总购卡数
    cpy_sht.cell(cpy_last_row, 3).value = sum_card_numb_now                                            # 所有企业的总绑卡量
    cpy_sht.cell(cpy_last_row, 4).value = sum_card_numb_now / sold_card_numb                           # 所有卡的投放总进度
    cpy_sht.cell(cpy_last_row, 7).value = sold_card_numb                                               # 所有企业的总购卡数
    cpy_sht.cell(cpy_last_row, 8).value = sum_card_numb_now / sold_card_numb                           # 所有卡的投放总进度
    cpy_sht.cell(cpy_last_row, 9).value = sum_card_numb_now - cpy_sht.cell(cpy_last_row, 11).value     # 本周新增绑卡数
    cpy_sht.cell(cpy_last_row, 10).value = sum_card_numb_now                                            # 所有企业的总绑卡量                      

    # ............《Chart_Data》............
    # 【0】绑卡进度
    #TODO:自动设置为选择绑卡进度最快的三个省份，且图表中省份名称是粘贴源自此的链接
    chart_data_sht['C2'].value = sum_card_numb_now                                              # 全国绑卡数
    chart_data_sht['C3'].value = prov_limit_dict['总计'][1] - sum_card_numb_now                 # 全国剩余量
    for i in [4, 6, 8]:                                                                         # 三省的绑卡数
        chart_data_sht.cell(i, 3).value = prov_dict[chart_data_sht.cell(i, 1).value]            
    for i in [5, 7, 9]:                                                                         # 三省的剩余量
        chart_data_sht.cell(i, 3).value = prov_limit_dict[chart_data_sht.cell(i, 1).value][1] - chart_data_sht.cell(i-1, 3).value   

    #【0】卡数量大字标
    chart_data_sht['F2'].value = sold_card_numb                             # 累计售卡数
    chart_data_sht['F3'].value = sum_card_numb_now                          # 累计绑卡数
    the_weekup_numb: int = prov_sht.cell(prov_last_row, 8).value
    chart_data_sht['F4'].value = the_weekup_numb                            # 本周新绑卡数
    chart_data_sht['F5'].value = avg_week_card                              # 平均周绑卡数

    chart_data_sht['I2'].value = f'学术卡数据周报 — {today_date}'    # 标题日期更新

    #【1】全国绑卡数趋势图
    for i in range(13, 24):
        chart_data_sht.cell(i,3).value = month_dict[str(chart_data_sht.cell(i,1).value)]

    #【2】TOP5 省份周增长绑卡数
    prov_weekup_dict: dict = {}      # 省份的本周增长数字典
    for i in range(2, prov_last_row):
        prov_weekup_dict.setdefault(prov_sht.cell(i,1).value, 0)
        prov_weekup_dict[prov_sht.cell(i, 1).value] += prov_sht.cell(i, 8).value

    # 生成按照新增数由大到小的元组列表
    sorted_prov_weekup_list: list = sorted(prov_weekup_dict.items(), key=lambda item:item[1],reverse=True)
    countr_avg: int = int(chart_data_sht['F4'].value / len(orig_prov_dict))     # 全国平均本周每个省的增长数量
    temp_index: int = 0                 # 排序列表中元组的索引
    for i in range(14,19):
        chart_data_sht.cell(i, 5).value = sorted_prov_weekup_list[temp_index][0]
        chart_data_sht.cell(i, 6).value = sorted_prov_weekup_list[temp_index][1]
        chart_data_sht.cell(i, 7).value = countr_avg
        temp_index += 1

    #【3】TOP5 卡类型周增长绑卡数箱型图
    card_weekup_dict: dict = {}
    for i in range(2, cpy_last_row):
        card_weekup_dict.setdefault(cpy_sht.cell(i,5).value,0)
        card_weekup_dict[cpy_sht.cell(i,5).value] += cpy_sht.cell(i,9).value
    sorted_card_weekup_list: list = sorted(card_weekup_dict.items(), key=lambda item:item[1],reverse=True)
    temp_index: int = 0                 # 排序列表中元组的索引
    other_card_weekup_numb: int = the_weekup_numb
    for i in range(14, 19):
        chart_data_sht.cell(i, 10).value = sorted_card_weekup_list[temp_index][0].replace('2019','').replace('2018','').replace('-5分','')
        chart_data_sht.cell(i, 11).value = sorted_card_weekup_list[temp_index][1]
        temp_index += 1
        other_card_weekup_numb = other_card_weekup_numb - sorted_card_weekup_list[temp_index][1]

    chart_data_sht['K19'].value = other_card_weekup_numb


    #【4】TOP10 省份绑卡数量分布图
    sorted_prov_list: list = sorted(prov_dict.items(), key=lambda item:item[1],reverse=True)
    temp_index: int = 0                 # 排序列表中元组的索引
    for i in range(27, 37):
        chart_data_sht.cell(i, 2).value = sorted_prov_list[temp_index][0]
        chart_data_sht.cell(i, 3).value = sorted_prov_list[temp_index][1]
        temp_index += 1


    #【5】省份绑卡状况数据地图
    for i in range(28, 62):
        chart_data_sht.cell(i,7).value = prov_dict[chart_data_sht.cell(i,6).value]

    #【6】企业购卡数量柱状图
    cpy_bought_dict: dict = {}      # 企业的本周增长数字典
    for k,v in cpy_dict.items():
        cpy_bought_dict.setdefault(k, 0)
        cpy_bought_dict[k] += v[0]

    # 生成按照新增数由大到小的元组列表
    sorted_cpy_bought_list: list = sorted(cpy_bought_dict.items(), key=lambda item:item[1],reverse=True)
    temp_index: int = 0                 # 排序列表中元组的索引
    other_bought_numb: int = sold_card_numb 
    for i in range(27, 34):
        chart_data_sht.cell(i, 10).value = sorted_cpy_bought_list[temp_index][0]
        chart_data_sht.cell(i, 11).value = sorted_cpy_bought_list[temp_index][1]
        temp_index += 1
        other_bought_numb = other_bought_numb - sorted_cpy_bought_list[temp_index][1]
    chart_data_sht.cell(34, 10).value = '其他企业'
    chart_data_sht.cell(34, 11).value = other_bought_numb

    #【7】TOP10 城市绑卡数量分布图
    sorted_city_list: list = sorted(city_dict.items(), key=lambda item:item[1],reverse=True)
    temp_index: int = 0                 # 排序列表中元组的索引
    for i in range(65, 75):
        chart_data_sht.cell(i, 2).value = sorted_city_list[temp_index][0]
        chart_data_sht.cell(i, 3).value = sorted_city_list[temp_index][1]
        temp_index += 1

    #【8】各小时内绑卡数量趋势图
    for i in range(65, 89):
        chart_data_sht.cell(i,7).value = hour_dict[str(chart_data_sht.cell(i,6).value)]

    #【9】绑卡医生医院级别比例
    for i in range(65, 72):
        chart_data_sht.cell(i,11).value = hosp_dict[str(chart_data_sht.cell(i,10).value)]


    # ------------------- 检查数据是否存在误差 -------------------
    #TODO:核对误差

    # ------------------- 表格的保存 -------------------
    data_wb.save(data_xlsx_path)
    history_path: str = os.path.join(workspace_path, 'history', today_date)
    if not os.path.exists(history_path):
        os.makedirs(history_path)
    report_wb_path: str = os.path.join(history_path, f'周报数据-{today_date}-UnMerged.xlsx')
    template_wb.save(report_wb_path)
    shutil.copy(os.path.join(workspace_path, '【图表模板】华医网学术卡数据周报.xlsx'),
                os.path.join(history_path, f'华医网学术卡数据周报-{today_date}.xlsx'))
    step_numb += 1
    print(f'【STEP-{step_numb}】文件保存\n\t\t[OK] --> 已经完成工作簿的保存！\n')


def convent_column_to_char(column_numb: int) -> str:
	"""
	【功能】将数字列数转换为Excel用英文字母表示的列数
	【示例】1 => A, 2 => B, ......, 27 => AA
	"""
	tStr: str = str()
	while column_numb != 0:
		res = column_numb % 26
		if res == 0:
			res = 26
			column_numb -= 26
		tStr = chr(ord('A') + res - 1) + tStr
		column_numb = column_numb // 26
	return tStr


def set_format():
    '''
    【功能】设置表格的样式
    '''
    global step_numb

    # 载入表格
    dst_xlsx_path: str = os.path.join(workspace_path, 'history', today_date, f'周报数据-{today_date}-UnMerged.xlsx')
    dst_wb = load_workbook(dst_xlsx_path)

    # 如果表格中未存在此定义样式，需要定义并声明
    try:
    # 公用样式
        grey_border = Border(left=Side(border_style='thin', color='c0c0c0'),
                            right=Side(border_style='thin', color='c0c0c0'),
                            top=Side(border_style='thin', color='c0c0c0'),
                            bottom=Side(border_style='thin', color='c0c0c0'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
        regular_font = Font(name='微软雅黑', size=11, bold=False, color='000000')

        # 设置首行的单元格样式
        title_style = NamedStyle(name='title_style', border=grey_border, alignment=center_align)
        title_style.font = Font(name='微软雅黑', size=11, bold=True, color='ffffff')
        title_style.fill = PatternFill("solid",fgColor='0070c0')

        # 设置常规内容的单元格样式
        white_content_style = NamedStyle(name='white_content_style', font=regular_font, border=grey_border, alignment=center_align,
                                        fill=PatternFill("solid",fgColor='ffffff'))
        blue_content_style = NamedStyle(name='blue_content_style', font=regular_font, border=grey_border, alignment=center_align,
                                        fill=PatternFill("solid",fgColor='d9e1f2'))

        # 设置总结行的单元格样式
        summary_style = NamedStyle(name='summary_style', alignment=center_align)
        summary_style.font = Font(name='微软雅黑', size=11, bold=True, color='0070c0')
        summary_style.border = Border(left=Side(border_style='thin', color='c0c0c0'),
                                    right=Side(border_style='thin',color='c0c0c0'),
                                    top=Side(border_style='medium',color='0070c0'),
                                    bottom=Side(border_style='thin', color='c0c0c0'))

        # 使用自定义样式之前需要在表格中声明
        dst_wb.add_named_style(title_style)
        dst_wb.add_named_style(white_content_style)
        dst_wb.add_named_style(blue_content_style)
        dst_wb.add_named_style(summary_style)
    except ValueError:
        pass

    sht_name_list: list = ['省份分布', '卡类状况', '企业投放统计']
    for sht_name in sht_name_list:
        cur_sht = dst_wb[sht_name]
        cur_last_col: int = cur_sht.max_column
        cur_last_row: int = cur_sht.max_row
        cur_last_row_char: str = convent_column_to_char(cur_last_col)
        # ----------- 设置单元格外观样式 --------------
        for single_cell in cur_sht[1]:
            single_cell.style = 'title_style'     # 设置首行的样式

        for single_cell in cur_sht[cur_last_row]:
            single_cell.style = 'summary_style'   # 设置总结行的样式

        first_col_value_list: list = []
        for i in range(2, cur_last_row):
            v: str = cur_sht.cell(i, 1).value
            if not v in first_col_value_list:
                first_col_value_list.append(v)

            if first_col_value_list.index(v) % 2 == 1:
                for single_cell in cur_sht[i]:
                    single_cell.style = 'blue_content_style'
            else:
                for single_cell in cur_sht[i]:
                    single_cell.style = 'white_content_style'

        # ----------- 设置单元格内容格式 --------------
        for j in range(1,11):  # 从前十列中寻找为“百分比格式”的列,设置显示样式
            if '进度' in cur_sht.cell(1,j).value or '比例' in cur_sht.cell(1,j).value:
                for x in range(2, cur_last_row + 1):
                    cur_sht.cell(x, j).number_format = '0.00%'


    dst_wb.save(dst_xlsx_path)
    step_numb += 1
    print(f'【STEP-{step_numb}】表格样式调整\n\t\t[OK] --> 已经表格样式的调整！\n')


def generate_new_template():
    '''
    【功能】生成新的数据模板
    '''
    global step_numb
    report_path: str = os.path.join(workspace_path, 'history', today_date, f'周报数据-{today_date}-UnMerged.xlsx')
    template_path: str = os.path.join(workspace_path, '【数据模板】华医网学术卡数据周报.xlsx')
    history_tempalte_path: str = os.path.join(workspace_path, 'template', f'Template_Data_{today_date}.xlsx')
    report_wb = load_workbook(report_path)

    # 清空《省份分布表》与《卡类状况表》数据
    for sheet_name in ['省份分布','卡类状况']:
        for i in range(2, report_wb[sheet_name].max_row + 1):
            for j in [2,3,4,5,6,8]:
                report_wb[sheet_name].cell(i,j).value = ''
        report_wb[sheet_name].cell(report_wb[sheet_name].max_row,2).value = '——'

    # 清空《企业投放统计表》数据
    for i in range(2, report_wb['企业投放统计'].max_row + 1):
        for j in [2,3,4,8,9]:
            report_wb['企业投放统计'].cell(i,j).value = ''
    
    # 清空《Chart_Data》数据
    data_rng_list: list = [ 'C2:C9', 'F2:F5', 'I2:I2', 
                            'C12:C23', 'E14:G18', 'J14:J18', 'K14:K19',
                            'B27:C36', 'G28:G61', 'J27:J33', 'K27:K34',
                            'B65:C74', 'G65:G88', 'K65:K71']
    for data_rng in data_rng_list:
        for cell_rng in report_wb['Chart_Data'][data_rng]:
            for j in range(len(cell_rng)):
                cell_rng[j].value = ''

    # 表格的保存
    shutil.move(template_path, history_tempalte_path)
    report_wb.save(template_path)

    step_numb += 1
    print(f'【STEP-{step_numb}】生成新模板\n\t\t[OK] --> 已经生成并保存新的数据模板！\n')


def merge_cells():
    '''
    【功能】根据首列单元格一致的，合并相应的单元格
    '''
    global step_numb

    src_xlsx_path: str = os.path.join(workspace_path, 'history', today_date, f'周报数据-{today_date}-UnMerged.xlsx')
    dst_xlsx_path: str = os.path.join(workspace_path, 'history', today_date, f'周报数据-{today_date}-Merged.xlsx')
    if not os.path.exists(dst_xlsx_path):
        shutil.copyfile(src_xlsx_path, dst_xlsx_path)
    dst_wb = load_workbook(dst_xlsx_path)
    
    sht_merge_col_dict: dict = {'省份分布':6, '卡类状况':6, '企业投放统计':4}
    for sht_name, merge_last_col in sht_merge_col_dict.items():
        cur_sht = dst_wb[sht_name]
        cur_last_row: int = cur_sht.max_row

        for i in range(cur_last_row-1, 1):
            if cur_sht.cell(i,1).value == cur_sht.cell(i-1,1).value:
                for j in range(1, merge_last_col+1):
                    col_char: str = convent_column_to_char(j)
                    cur_sht.merge_cells(f'{col_char}{str(i-1)}:{col_char}{str(i)}')

    dst_wb.save(dst_xlsx_path)
    step_numb += 1
    print(f'【STEP-{step_numb}】合并单元格\n\t\t[OK] --> 已经根据首列合并单元格并保存文件！\n')


# ------------------------------ 主体调用部分 ------------------------------
def huayi_card_report():
    data_result: list = download_data_xls()
    global step_numb
    if data_result[0] == 200:
        step_numb += 1
        print(f'【STEP-{step_numb}】爬虫下载原始数据\n\t\t[OK] --> 已经成功下载数据文件!\n')
        xlsx_path: str = xls_to_xlsx(data_result[1])
        step_numb += 1
        print(f'【STEP-{step_numb}】文件格式转换\n\t\t[OK] --> 已经将文件转化为xlsx格式!\n')
        statistic_data()
        set_format()
        generate_new_template()
        merge_cells()
        print('-'*100 + '\n统计完成，请自行检查核对一下数据及样式是否正确！！！\n')
    else:
        step_numb += 1
        print(f'【STEP-{step_numb}】爬虫下载原始数据\n\t\t[ERROR] --> 未能从服务器中成功爬取数据:状态码为 {data_result[0]}\n')


huayi_card_report()

#TODO: 进度预警
#TODO:周增长数为负数且大于等于5，提醒
#BUG:有其他的两个chartdata表的其他有问题
